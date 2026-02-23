import random
from uuid import uuid4

from app import create_app
from app.auth import hash_password
from app.database import init_db, get_db, query_db


COVER_COLORS = [
    '#4A90E2', '#E74C3C', '#2ECC71', '#F39C12', '#9B59B6',
    '#1ABC9C', '#E67E22', '#3498DB', '#E91E63', '#00BCD4',
    '#8BC34A', '#FF5722',
]

BOOKS = [
    {'title': 'Котигорошко', 'category': 'Казки', 'author': 'Народна творчість'},
    {'title': 'Кирпатий казкар', 'category': 'Казки', 'author': 'Іван Франко'},
    {'title': 'Лісова пісня', 'category': 'Казки', 'author': 'Леся Українка'},
    {'title': 'Тореадори з Васюківки', 'category': 'Пригоди', 'author': 'Всеволод Нестайко'},
    {'title': 'Пригоди Барвінка', 'category': 'Пригоди', 'author': 'Богдан Чалий'},
    {'title': 'Микита Кожум\'яка', 'category': 'Казки', 'author': 'Народна творчість'},
    {'title': 'Чомусик і Томусик', 'category': 'Наука', 'author': 'Всеволод Нестайко'},
    {'title': 'Кайдашева сім\'я', 'category': 'Казки', 'author': 'Іван Нечуй-Левицький'},
    {'title': 'Маруся Чурай', 'category': 'Пригоди', 'author': 'Ліна Костенко'},
    {'title': 'Захар Беркут', 'category': 'Пригоди', 'author': 'Іван Франко'},
    {'title': 'Фарбований Лис', 'category': 'Казки', 'author': 'Іван Франко'},
    {'title': 'Рудко', 'category': 'Пригоди', 'author': 'Олесь Донченко'},
]


def seed():
    """Seed the database with initial data.

    Usage:
        python -c "from app.seed import seed; seed()"
    """
    app = create_app()

    with app.app_context():
        init_db()

        # Check if admin already exists
        existing_admin = query_db(
            'SELECT id FROM users WHERE email = ?',
            ['admin@library.com'],
            one=True
        )
        if existing_admin:
            print('Admin user already exists. Skipping seed.')
            return

        db = get_db()

        # --- Create admin user ---
        admin_id = str(uuid4())
        db.execute(
            'INSERT INTO users (id, email, password_hash, full_name, role) VALUES (?, ?, ?, ?, ?)',
            [admin_id, 'admin@library.com', hash_password('admin123'), 'Адміністратор', 'admin']
        )
        print('Created admin user: admin@library.com / admin123')

        # --- Create categories ---
        category_ids = {}
        for name in ['Казки', 'Пригоди', 'Наука']:
            cat_id = str(uuid4())
            db.execute('INSERT INTO categories (id, name) VALUES (?, ?)', [cat_id, name])
            category_ids[name] = cat_id
        print('Created 3 categories.')

        # --- Create publishers ---
        publishers = [
            ('А-БА-БА-ГА-ЛА-МА-ГА', 'Київ'),
            ('Видавництво Старого Лева', 'Львів'),
        ]
        publisher_ids = []
        for pub_name, pub_city in publishers:
            pub_id = str(uuid4())
            db.execute(
                'INSERT INTO publishers (id, name, city) VALUES (?, ?, ?)',
                [pub_id, pub_name, pub_city]
            )
            publisher_ids.append(pub_id)
        print('Created 2 publishers.')

        # --- Create series ---
        series_list = ['Казки народів світу', 'Маленький дослідник']
        series_ids = []
        for s_name in series_list:
            s_id = str(uuid4())
            db.execute('INSERT INTO series (id, name) VALUES (?, ?)', [s_id, s_name])
            series_ids.append(s_id)
        print('Created 2 series.')

        # --- Create 12 books ---
        shuffled_colors = COVER_COLORS[:]
        random.shuffle(shuffled_colors)

        for idx, book_data in enumerate(BOOKS):
            book_id = str(uuid4())
            cat_id = category_ids[book_data['category']]
            color = shuffled_colors[idx % len(shuffled_colors)]

            db.execute(
                '''INSERT INTO books (id, title, author, category, category_id, cover_color, available, new_book)
                   VALUES (?, ?, ?, ?, ?, ?, 1, 0)''',
                [book_id, book_data['title'], book_data['author'],
                 book_data['category'], cat_id, color]
            )
        print('Created 12 books.')

        db.commit()
        print('Seed completed successfully.')


def seed_from_excel(filepath):
    """Seed the database with books from an Excel file.

    Clears all book-related data (books, categories, series, publishers, book_media, rental_requests)
    and repopulates from the "КНИГИ" sheet.

    Usage:
        python -c "from app.seed import seed_from_excel; seed_from_excel('../Бібліотечка Українського дитячого клубу.xlsx')"
    """
    import openpyxl

    app = create_app()

    with app.app_context():
        init_db()
        db = get_db()

        # --- Clear book-related tables (FK order) ---
        db.execute('DELETE FROM book_media')
        db.execute('DELETE FROM rental_requests')
        db.execute('DELETE FROM books')
        db.execute('DELETE FROM categories')
        db.execute('DELETE FROM series')
        db.execute('DELETE FROM publishers')
        print('Cleared existing book-related data.')

        # --- Read Excel ---
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb['КНИГИ']

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
        wb.close()
        print(f'Read {len(rows)} rows from Excel.')

        # --- First pass: collect unique categories, publishers, series ---
        categories_set = set()
        publishers_dict = {}  # name -> city
        series_set = set()

        for row in rows:
            title = row[5]
            author = row[4]
            if not title or not str(title).strip():
                continue

            cat = row[3]
            if cat and str(cat).strip():
                categories_set.add(str(cat).strip())

            pub_name = row[9]
            pub_city = row[8]
            if pub_name and str(pub_name).strip():
                pub_name_clean = str(pub_name).strip()
                pub_city_clean = str(pub_city).strip() if pub_city else ''
                if pub_name_clean not in publishers_dict:
                    publishers_dict[pub_name_clean] = pub_city_clean

            ser = row[7]
            if ser and str(ser).strip():
                series_set.add(str(ser).strip())

        # --- Insert categories ---
        category_ids = {}
        for name in sorted(categories_set):
            cat_id = str(uuid4())
            db.execute('INSERT INTO categories (id, name) VALUES (?, ?)', [cat_id, name])
            category_ids[name] = cat_id
        print(f'Created {len(category_ids)} categories.')

        # --- Insert publishers ---
        publisher_ids = {}
        for name in sorted(publishers_dict.keys()):
            pub_id = str(uuid4())
            city = publishers_dict[name]
            db.execute('INSERT INTO publishers (id, name, city) VALUES (?, ?, ?)', [pub_id, name, city])
            publisher_ids[name] = pub_id
        print(f'Created {len(publisher_ids)} publishers.')

        # --- Insert series ---
        series_ids = {}
        for name in sorted(series_set):
            s_id = str(uuid4())
            db.execute('INSERT INTO series (id, name) VALUES (?, ?)', [s_id, name])
            series_ids[name] = s_id
        print(f'Created {len(series_ids)} series.')

        # --- Insert books ---
        colors = COVER_COLORS[:]
        book_count = 0
        skipped = 0

        for row in rows:
            title = row[5]
            author = row[4]
            if not title or not str(title).strip():
                skipped += 1
                continue

            title = str(title).strip()
            author = str(author).strip() if author else ''

            book_id = str(uuid4())
            color = random.choice(colors)

            # Category
            cat_name = str(row[3]).strip() if row[3] else None
            cat_id = category_ids.get(cat_name) if cat_name else None

            # Inventory number
            inv = row[2]
            inventory_number = None
            if inv is not None:
                try:
                    inventory_number = int(float(inv))
                except (ValueError, TypeError):
                    pass

            # Series
            ser = str(row[7]).strip() if row[7] else None
            series_id = series_ids.get(ser) if ser else None

            # Publisher
            pub = str(row[9]).strip() if row[9] else None
            publisher_id = publisher_ids.get(pub) if pub else None

            # Publication year
            year_val = row[10]
            publication_year = None
            if year_val is not None:
                try:
                    publication_year = str(int(float(year_val)))
                except (ValueError, TypeError):
                    publication_year = str(year_val).strip() if str(year_val).strip() else None

            # ISBN
            isbn = str(row[11]).strip() if row[11] else None

            # Supplier
            supplier = str(row[12]).strip() if row[12] else None

            # Age
            age = str(row[14]).strip() if row[14] else None

            # Description
            description = str(row[15]).strip() if row[15] else None

            # Available — col 6 (Наявність) is more complete than col 18
            avail_val = row[6]
            available = 0 if avail_val and str(avail_val).strip() == 'ЧИТАЮТЬ' else 1

            db.execute(
                '''INSERT INTO books (id, title, author, category, category_id, series_id, publisher_id,
                   cover_color, available, description, age, publication_year, isbn,
                   inventory_number, supplier, new_book)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)''',
                [book_id, title, author, cat_name or '', cat_id, series_id, publisher_id,
                 color, available, description, age, publication_year, isbn,
                 inventory_number, supplier]
            )
            book_count += 1

        db.commit()
        print(f'Created {book_count} books. Skipped {skipped} rows (missing title/author).')

        # --- Create approved rental requests from "Книги на руках QUERY" sheet ---
        wb2 = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws2 = wb2['Книги на руках QUERY']

        # Build case-insensitive title → (book_id, book_title) lookup
        all_books = db.execute('SELECT id, title FROM books').fetchall()
        title_lookup = {}  # lowercase title → (id, original_title)
        for b in all_books:
            title_lookup[b[1].lower()] = (b[0], b[1])

        rental_count = 0
        unmatched = []

        for row in ws2.iter_rows(min_row=1, values_only=True):
            book_title_raw = row[1]
            renter_raw = row[2]
            rental_date = row[0]
            if not book_title_raw or not str(book_title_raw).strip():
                continue

            book_title_clean = str(book_title_raw).strip()
            renter_name = str(renter_raw).strip() if renter_raw else 'Невідомий'

            match = title_lookup.get(book_title_clean.lower())
            if not match:
                unmatched.append(book_title_clean)
                continue

            book_id, original_title = match

            # Format approved_at from the date column
            approved_at = None
            if rental_date and hasattr(rental_date, 'strftime'):
                approved_at = rental_date.strftime('%Y-%m-%d %H:%M:%S')

            req_id = str(uuid4())
            db.execute(
                '''INSERT INTO rental_requests
                   (id, book_id, book_title, renter_name, renter_phone, renter_email,
                    rental_duration, status, requested_at, approved_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, 'approved', ?, ?)''',
                [req_id, book_id, original_title, renter_name,
                 '7777777', 'blank@gmail.com', 4,
                 approved_at or '2025-01-01 00:00:00', approved_at]
            )

            # Ensure book is marked unavailable
            db.execute('UPDATE books SET available = 0 WHERE id = ?', [book_id])
            rental_count += 1

        wb2.close()
        db.commit()

        print(f'Created {rental_count} approved rental requests.')
        if unmatched:
            print(f'  Could not match {len(unmatched)} titles:')
            for t in unmatched:
                print(f'    - {t}')

        print('Excel seed completed successfully.')


if __name__ == '__main__':
    seed()
