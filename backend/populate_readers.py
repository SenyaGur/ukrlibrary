"""
Populate readers and children from the Excel spreadsheet.
Link existing rental_requests to the correct reader_id.
"""
import sqlite3
import uuid
import openpyxl
import re
import sys

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else '../2026-02-21 Копия Бібліотечка Українського дитячого клубу.xlsx'
DB_PATH = sys.argv[2] if len(sys.argv) > 2 else 'library.db'

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['ЧИТАЧІ']

# --- Parse parents ---
parents = {}  # id_parent -> {name, phone1, phone2, address, comment}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    full_name = row[0]
    id_parent = row[1]
    phone_ukr = row[2]
    phone_bg = row[3]
    district = row[4]
    comment = row[5]

    if full_name is None or id_parent is None:
        continue

    full_name = str(full_name).strip()
    id_parent = int(id_parent)

    # Parse phones - handle formula strings like =359884206203, ="...", plain numbers
    def clean_phone(val):
        if val is None:
            return ''
        s = str(val).strip()
        # Remove formula markers
        s = s.replace('=', '').replace('"', '').replace("'", '')
        # Remove non-breaking spaces, special chars
        s = re.sub(r'[\s\u00a0\u202c]+', '', s)
        if not s or s == 'None' or s == '0':
            return ''
        # If it's a pure number, format it
        try:
            n = int(float(s))
            s = str(n)
        except (ValueError, OverflowError):
            pass
        # Add + prefix if it looks like an international number without one
        if len(s) >= 10 and s[0].isdigit():
            s = '+' + s
        return s

    p1 = clean_phone(phone_ukr)
    p2 = clean_phone(phone_bg)

    # Use whichever phone is available as phone1
    if not p1 and p2:
        p1, p2 = p2, ''

    address = str(district).strip() if district else ''
    if address == 'None':
        address = ''

    parents[id_parent] = {
        'full_name': full_name,
        'phone1': p1 if p1 else 'не вказано',
        'phone2': p2 if p2 else None,
        'address': address if address else 'не вказано',
        'comment': str(comment).strip() if comment and str(comment).strip() != 'None' else None,
    }

# --- Parse children ---
children_data = []  # list of {parent_id, name, birthday}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    kid_id = row[8]
    parent_id = row[9]
    kid_name = row[10]
    kid_birthday = row[11]
    kid_gender = row[12]

    if kid_name is None or parent_id is None:
        continue

    parent_id = int(parent_id)
    kid_name = str(kid_name).strip()

    bday = ''
    if kid_birthday:
        try:
            bday = kid_birthday.strftime('%Y-%m-%d')
        except:
            bday = str(kid_birthday)

    children_data.append({
        'parent_id': parent_id,
        'name': kid_name,
        'birthday': bday if bday else '2000-01-01',
    })

print(f"Parsed {len(parents)} parents and {len(children_data)} children from Excel")

# --- Insert into DB ---
db = sqlite3.connect(DB_PATH)
db.row_factory = sqlite3.Row

# Check existing readers
existing = db.execute("SELECT COUNT(*) as c FROM readers").fetchone()['c']
if existing > 0:
    print(f"WARNING: {existing} readers already exist. Clearing them first.")
    db.execute("DELETE FROM children")
    db.execute("DELETE FROM readers")
    db.commit()

# Map parent excel id -> db uuid
parent_uuid_map = {}

for pid, p in sorted(parents.items()):
    full = p['full_name']
    # Split into name and surname
    parts = full.split(None, 1)
    if len(parts) == 2:
        # Could be "Name Surname" or "Surname Name"
        # The Excel seems to have "Name Surname" format in most cases
        first_name = parts[0]
        surname = parts[1]
    else:
        first_name = full
        surname = ''

    reader_id = str(uuid.uuid4())
    parent_uuid_map[pid] = reader_id

    db.execute(
        "INSERT INTO readers (id, parent_name, parent_surname, phone1, phone2, address) VALUES (?, ?, ?, ?, ?, ?)",
        (reader_id, first_name, surname, p['phone1'], p['phone2'], p['address'])
    )

print(f"Inserted {len(parent_uuid_map)} readers")

# Insert children
children_inserted = 0
for child in children_data:
    pid = child['parent_id']
    reader_id = parent_uuid_map.get(pid)
    if not reader_id:
        # parent_id in kids doesn't match any parent row - try to find
        # Some kids have parent_id that doesn't match the row number
        continue

    child_id = str(uuid.uuid4())
    # Use parent's surname for child
    parent_info = parents.get(pid, {})
    parent_full = parent_info.get('full_name', '')
    parts = parent_full.split(None, 1)
    child_surname = parts[1] if len(parts) == 2 else parts[0] if parts else ''

    db.execute(
        "INSERT INTO children (id, reader_id, name, surname, birth_date) VALUES (?, ?, ?, ?, ?)",
        (child_id, reader_id, child['name'], child_surname, child['birthday'])
    )
    children_inserted += 1

print(f"Inserted {children_inserted} children")

# --- Link rental_requests to readers ---
# Get all distinct renter names from rentals
rentals = db.execute("SELECT DISTINCT renter_name FROM rental_requests").fetchall()
renter_names = [r['renter_name'] for r in rentals]

# Build mapping: renter_name -> reader_id
# Match by checking if renter_name starts with or contains the parent's full name
linked = 0
for renter_name in renter_names:
    rn_lower = renter_name.lower().strip()
    best_match = None
    best_len = 0

    for pid, p in parents.items():
        pname = p['full_name'].lower().strip()
        # Check if renter_name starts with parent name
        if rn_lower.startswith(pname) and len(pname) > best_len:
            best_match = pid
            best_len = len(pname)
        # Also check exact match
        elif rn_lower == pname and len(pname) > best_len:
            best_match = pid
            best_len = len(pname)

    if best_match:
        reader_id = parent_uuid_map[best_match]
        count = db.execute(
            "UPDATE rental_requests SET reader_id = ? WHERE renter_name = ?",
            (reader_id, renter_name)
        ).rowcount
        linked += count
        print(f"  Linked '{renter_name}' -> {parents[best_match]['full_name']} ({count} requests)")
    else:
        print(f"  NO MATCH for '{renter_name}'")

print(f"\nLinked {linked} rental requests to readers")

db.commit()
db.close()
print("Done!")
